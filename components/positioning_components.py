"""스펙 포지셔닝 분석 UI 컴포넌트."""

from __future__ import annotations

import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 컬럼 매핑 UI
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def render_column_mapping_ui(
    raw_df: pd.DataFrame,
    auto_config: dict,
) -> dict:
    """컬럼 자동 감지 결과를 사용자가 확인/수정할 수 있는 UI."""
    all_cols = raw_df.columns.tolist()

    col1, col2 = st.columns(2)

    with col1:
        product_idx = (
            all_cols.index(auto_config["product_col"])
            if auto_config.get("product_col") in all_cols
            else 0
        )
        product_col = st.selectbox(
            "제품명 컬럼", all_cols, index=product_idx, key="spec_product_col",
        )

        price_idx = (
            all_cols.index(auto_config["price_col"])
            if auto_config.get("price_col") in all_cols
            else min(1, len(all_cols) - 1)
        )
        price_col = st.selectbox(
            "가격 컬럼", all_cols, index=price_idx, key="spec_price_col",
        )

    with col2:
        cat_options = ["(없음)"] + all_cols
        cat_idx = (
            cat_options.index(auto_config["category_col"])
            if auto_config.get("category_col") in cat_options
            else 0
        )
        category_col = st.selectbox(
            "카테고리 컬럼 (선택)", cat_options, index=cat_idx, key="spec_category_col",
        )
        if category_col == "(없음)":
            category_col = None

    # 스펙 컬럼: 나머지 전부
    excluded = {product_col, price_col}
    if category_col:
        excluded.add(category_col)
    remaining = [c for c in all_cols if c not in excluded]

    default_specs = [c for c in auto_config.get("spec_cols", []) if c in remaining]

    spec_cols = st.multiselect(
        "스펙 컬럼 선택 (분석 대상)",
        remaining,
        default=default_specs or remaining,
        key="spec_spec_cols",
    )

    if not spec_cols:
        st.warning("스펙 컬럼을 1개 이상 선택해주세요.")

    return {
        "product_col": product_col,
        "price_col": price_col,
        "category_col": category_col,
        "spec_cols": spec_cols,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 가중치 슬라이더
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def render_weight_sliders(
    spec_cols: list[str],
    default_weights: dict[str, float],
) -> dict[str, float]:
    """각 스펙 컬럼의 가중치를 슬라이더로 조정하는 UI. 합계 1.0 자동 정규화."""
    raw_weights: dict[str, float] = {}
    cols_per_row = 3

    for i in range(0, len(spec_cols), cols_per_row):
        chunk = spec_cols[i:i + cols_per_row]
        cols = st.columns(len(chunk))
        for j, sc in enumerate(chunk):
            with cols[j]:
                val = st.slider(
                    sc,
                    min_value=0.0, max_value=1.0,
                    value=float(default_weights.get(sc, 0.5)),
                    step=0.05,
                    key=f"weight_{sc}",
                )
                raw_weights[sc] = val

    total = sum(raw_weights.values())
    if total > 0:
        normalized = {k: v / total for k, v in raw_weights.items()}
    else:
        n = len(raw_weights) or 1
        normalized = {k: 1.0 / n for k in raw_weights}

    st.caption(f"가중치 합계: 1.00 (자동 정규화 적용)")
    return normalized


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 데이터 테이블
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def render_scored_data_table(
    scored_df: pd.DataFrame,
    column_config: dict,
    categories: dict[str, str] | None = None,
) -> None:
    """점수화된 데이터 테이블을 렌더링한다."""
    display_cols = [
        column_config["product_col"],
        column_config["price_col"],
    ] + column_config["spec_cols"] + ["spec_score"]

    display_df = scored_df[display_cols].copy()
    display_df["spec_score"] = display_df["spec_score"].round(1)

    # 카테고리 컬럼 추가
    if categories:
        display_df.insert(
            1, "포지셔닝",
            display_df[column_config["product_col"]].map(
                lambda x: categories.get(str(x), "")
            ),
        )

    display_df = display_df.sort_values("spec_score", ascending=False)

    st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "spec_score": st.column_config.ProgressColumn(
                "스펙 점수", format="%.1f", min_value=0, max_value=100,
            ),
        },
    )
    st.caption(f"총 {len(display_df)}개 제품")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 시뮬레이션 폼
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def render_simulation_form(
    column_config: dict,
    raw_df: pd.DataFrame,
) -> dict | None:
    """우리 제품 시뮬레이션 입력 폼을 렌더링한다."""
    with st.form("simulation_form"):
        st.markdown("**우리 제품의 예상 스펙과 가격을 입력하세요**")

        form_c1, form_c2 = st.columns(2)
        with form_c1:
            product_name = st.text_input("제품명", value="우리 제품")
        with form_c2:
            price_col = column_config["price_col"]
            median_price = int(raw_df[price_col].median())
            price = st.number_input(
                "가격", min_value=0, value=median_price, step=10000,
            )

        specs: dict = {}
        cols_per_row = 3
        spec_cols = column_config["spec_cols"]

        for i in range(0, len(spec_cols), cols_per_row):
            chunk = spec_cols[i:i + cols_per_row]
            cols = st.columns(len(chunk))
            for j, sc in enumerate(chunk):
                with cols[j]:
                    col_data = raw_df[sc]
                    if pd.api.types.is_numeric_dtype(col_data):
                        val = st.number_input(
                            sc,
                            value=float(col_data.median()),
                            key=f"sim_{sc}",
                        )
                    else:
                        unique_vals = col_data.dropna().unique().tolist()
                        val = st.selectbox(sc, unique_vals, key=f"sim_{sc}")
                    specs[sc] = val

        submitted = st.form_submit_button("시뮬레이션 실행", type="primary")

    if submitted:
        result = {"product_name": product_name, "price": price}
        result.update(specs)
        return result
    return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# AI 분석 결과 렌더링
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def render_ai_analysis(analysis: dict) -> None:
    """AI 전략 분석 결과를 렌더링한다."""
    # 시장 개요
    overview = analysis.get("market_overview", "")
    if overview:
        st.info(f"**시장 개요**\n\n{overview}")

    col_a, col_b = st.columns(2)

    # 과밀 영역
    with col_a:
        overcrowded = analysis.get("overcrowded_zones", [])
        if overcrowded:
            st.markdown("**과밀 경쟁 영역**")
            for zone in overcrowded:
                st.markdown(f"- :red[{zone}]")

    # 공백 영역
    with col_b:
        gap_areas = analysis.get("gap_areas", [])
        if gap_areas:
            st.markdown("**시장 공백 / 기회 영역**")
            for gap in gap_areas:
                st.markdown(f"- :green[{gap}]")

    # 가치 지수 분석
    value_analysis = analysis.get("value_index_analysis", "")
    if value_analysis:
        with st.expander("가치 지수 분석", expanded=True):
            st.markdown(value_analysis)

    # 전략 권고
    recommendations = analysis.get("recommendations", [])
    if recommendations:
        st.markdown("---")
        st.markdown("**전략 권고**")
        for i, rec in enumerate(recommendations, 1):
            st.success(f"**권고 {i}.** {rec}")

    # 우리 제품 평가
    our_assessment = analysis.get("our_product_assessment", "")
    if our_assessment:
        st.markdown("---")
        st.markdown("**우리 제품 포지셔닝 평가**")
        st.warning(our_assessment)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 내보내기
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def export_positioning_to_excel(
    scored_df: pd.DataFrame,
    weights: dict[str, float],
    column_config: dict,
) -> bytes:
    """분석 결과를 엑셀 파일로 내보낸다 (2시트: 분석결과 + 가중치)."""
    wb = Workbook()

    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    body_font = Font(name="맑은 고딕", size=10)

    # 시트 1: 포지셔닝 분석
    ws1 = wb.active
    ws1.title = "포지셔닝 분석"

    display_cols = (
        [column_config["product_col"], column_config["price_col"]]
        + column_config["spec_cols"]
        + ["spec_score"]
    )

    for j, col in enumerate(display_cols, 1):
        cell = ws1.cell(row=1, column=j, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, (_, row) in enumerate(scored_df.iterrows(), 2):
        for j, col in enumerate(display_cols, 1):
            val = row.get(col)
            if col == "spec_score":
                val = round(val, 1)
            cell = ws1.cell(row=i, column=j, value=val)
            cell.font = body_font

    # 컬럼 너비 자동 조정
    for j, col in enumerate(display_cols, 1):
        ws1.column_dimensions[chr(64 + j) if j <= 26 else "A"].width = max(12, len(str(col)) + 4)

    # 시트 2: 가중치
    ws2 = wb.create_sheet("가중치")
    for j, header in enumerate(["스펙 항목", "가중치"], 1):
        cell = ws2.cell(row=1, column=j, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for i, (name, weight) in enumerate(
        sorted(weights.items(), key=lambda x: x[1], reverse=True), 2
    ):
        ws2.cell(row=i, column=1, value=name).font = body_font
        cell = ws2.cell(row=i, column=2, value=weight)
        cell.font = body_font
        cell.number_format = "0.0%"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_strategy_report(analysis: dict) -> str:
    """AI 분석 결과를 텍스트 보고서로 포맷팅한다."""
    lines = [
        "=" * 60,
        "스펙 기반 포지셔닝 전략 보고서",
        "=" * 60,
        "",
        "[시장 개요]",
        analysis.get("market_overview", ""),
        "",
        "[밀집 영역 (과밀 경쟁)]",
    ]
    for z in analysis.get("overcrowded_zones", []):
        lines.append(f"  - {z}")

    lines += ["", "[공백 영역 / 기회]"]
    for g in analysis.get("gap_areas", []):
        lines.append(f"  - {g}")

    lines += ["", "[가치 지수 분석]", analysis.get("value_index_analysis", ""), ""]

    lines.append("[전략 권고]")
    for i, r in enumerate(analysis.get("recommendations", []), 1):
        lines.append(f"  {i}. {r}")

    our = analysis.get("our_product_assessment", "")
    if our:
        lines += ["", "[우리 제품 포지셔닝 평가]", our]

    return "\n".join(lines)
