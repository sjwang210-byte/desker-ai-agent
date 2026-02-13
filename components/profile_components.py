"""고객 프로파일 분석 UI 컴포넌트."""

from __future__ import annotations

import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def render_upload_status(profile_data: dict) -> None:
    """3개 차원의 업로드 상태 카드를 렌더링한다."""
    all_dimensions = ["자녀나이", "결혼상태", "가구당인원"]
    cols = st.columns(3)

    for i, dim in enumerate(all_dimensions):
        with cols[i]:
            if dim in profile_data:
                data = profile_data[dim]
                st.success(f"**{dim}** — 업로드 완료")
                st.caption(f"파일: {data['filename']}")
                n_products = data["df"]["attribute_value"].nunique()
                n_rows = len(data["df"])
                st.caption(f"행 수: {n_rows:,}개 (속성값 {n_products}종)")
                dr = data.get("date_range", {})
                if dr:
                    st.caption(f"기간: {dr.get('start', '?')} ~ {dr.get('end', '?')}")
            else:
                st.warning(f"**{dim}** — 미업로드")


def render_percentage_table(
    result_df: pd.DataFrame,
    attribute_values: list[str],
    show_absolute: bool = False,
) -> None:
    """비중(%) 분석 결과 테이블을 렌더링한다."""
    if result_df.empty:
        st.info("분석 결과가 없습니다.")
        return

    display_df = result_df[["category"]].copy()

    for attr in attribute_values:
        if attr not in result_df.columns:
            continue
        if show_absolute:
            abs_col = f"{attr}_abs"
            if abs_col in result_df.columns:
                display_df[attr] = result_df.apply(
                    lambda r, a=attr, ac=abs_col: f"{r[a]:.1f}% ({r[ac]:,.0f})",
                    axis=1,
                )
            else:
                display_df[attr] = result_df[attr].apply(lambda x: f"{x:.1f}%")
        else:
            display_df[attr] = result_df[attr].apply(lambda x: f"{x:.1f}%")

    if "합계" in result_df.columns:
        display_df["합계"] = result_df["합계"].apply(lambda x: f"{x:,.0f}")

    display_df = display_df.rename(columns={"category": "카테고리"})
    st.dataframe(display_df, use_container_width=True, hide_index=True)
    st.caption(f"총 {len(display_df)}개 카테고리")


def render_drilldown_selector(
    result_df: pd.DataFrame,
    current_level: str,
    child_level: str | None,
) -> str | None:
    """드릴다운 카테고리 선택 위젯을 렌더링한다.

    Returns:
        선택된 카테고리 값 또는 None (최하위 레벨).
    """
    if child_level is None:
        st.info("최하위 수준입니다. 더 이상 드릴다운할 수 없습니다.")
        return None

    if result_df.empty:
        return None

    categories = result_df["category"].tolist()
    selected = st.selectbox(
        f"드릴다운: {current_level} → {child_level}",
        categories,
        key=f"drilldown_{current_level}",
    )
    return selected


def render_integrated_view(
    integrated_results: dict[str, pd.DataFrame],
    category_value: str,
    metric: str,
) -> None:
    """3개 차원을 나란히 메트릭으로 표시한다."""
    if not integrated_results:
        st.info("통합 분석 데이터가 없습니다.")
        return

    st.markdown(f"**'{category_value}'** 고객 프로파일 ({metric} 기준)")

    cols = st.columns(len(integrated_results))
    for i, (dimension, df) in enumerate(integrated_results.items()):
        with cols[i]:
            st.markdown(f"##### {dimension}")
            if df.empty:
                st.info("데이터 없음")
                continue

            attr_cols = [
                c for c in df.columns
                if not c.endswith("_abs") and c != "합계"
            ]
            for attr in attr_cols:
                pct = df[attr].iloc[0]
                abs_col = f"{attr}_abs"
                abs_val = df[abs_col].iloc[0] if abs_col in df.columns else None
                delta_str = f"{abs_val:,.0f}" if abs_val is not None else None
                st.metric(attr, f"{pct:.1f}%", delta=delta_str, delta_color="off")


def export_profile_to_excel(
    result_df: pd.DataFrame,
    dimension: str,
    metric: str,
    agg_level: str,
    attribute_values: list[str],
) -> bytes:
    """분석 결과를 엑셀 파일(bytes)로 내보낸다."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{dimension}_{agg_level}"

    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(
        start_color="1E88E5", end_color="1E88E5", fill_type="solid",
    )
    body_font = Font(name="맑은 고딕", size=10)
    center = Alignment(horizontal="center")

    # 제목
    title_cell = ws.cell(
        row=1, column=1,
        value=f"고객 프로파일 분석: {dimension} / {metric} / {agg_level}",
    )
    title_cell.font = Font(name="맑은 고딕", bold=True, size=12)

    # 헤더 (3행)
    headers = ["카테고리"] + attribute_values + ["합계"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 데이터
    for i, (_, row) in enumerate(result_df.iterrows(), start=4):
        ws.cell(row=i, column=1, value=row["category"]).font = body_font
        for j, attr in enumerate(attribute_values, start=2):
            pct_val = row.get(attr, 0)
            cell = ws.cell(row=i, column=j, value=pct_val / 100)
            cell.number_format = '0.0"%"'
            cell.font = body_font
            cell.alignment = center

        total_cell = ws.cell(
            row=i, column=len(attribute_values) + 2, value=row.get("합계", 0),
        )
        total_cell.number_format = "#,##0"
        total_cell.font = body_font

    # 열 너비
    ws.column_dimensions["A"].width = 25
    for j in range(2, len(headers) + 1):
        ws.column_dimensions[chr(64 + j)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
