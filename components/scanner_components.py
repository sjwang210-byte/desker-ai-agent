"""시장 스캐너 UI 컴포넌트."""

import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def render_product_card(product: dict, index: int):
    """단일 제품 정보 카드 렌더링."""
    has_error = product.get("error", False)

    with st.container(border=True):
        if has_error:
            st.warning(f"**#{index}** {product.get('error_message', 'AI 분석 실패')}")

        col_img, col_info = st.columns([1, 3])

        with col_img:
            image_url = product.get("image_url", "")
            if image_url:
                st.image(image_url, width=200)
            else:
                st.markdown("*이미지 없음*")

        with col_info:
            st.subheader(f"#{index} {product.get('product_name', '정보 없음')}")

            info_col1, info_col2 = st.columns(2)
            with info_col1:
                st.markdown(f"**브랜드:** {product.get('brand', '정보 없음')}")
                st.markdown(f"**가격:** {product.get('price_display', '정보 없음')}")
                st.markdown(f"**원산지:** {product.get('country_of_origin', '정보 없음')}")
                st.markdown(f"**크기:** {product.get('size', '정보 없음')}")

            with info_col2:
                st.markdown(f"**소재:** {product.get('materials', '정보 없음')}")
                options = product.get("options", [])
                if options:
                    st.markdown(f"**옵션:** {', '.join(options)}")
                else:
                    st.markdown("**옵션:** 정보 없음")

                features = product.get("notable_features", [])
                if features:
                    st.markdown("**특이사항:**")
                    for f in features:
                        st.markdown(f"- {f}")

        # 리뷰 요약
        review = product.get("review_summary", {})
        if review and review.get("summary_text"):
            _render_review(review)


def _render_review(review: dict):
    """리뷰 분석 결과 렌더링."""
    st.markdown("---")
    review_cols = st.columns([1, 1, 2])

    with review_cols[0]:
        count = review.get("total_count", 0)
        rating = review.get("average_rating", 0)
        if count:
            st.metric("리뷰 수", f"{count:,}건")
        if rating:
            stars = "★" * int(rating) + "☆" * (5 - int(rating))
            st.metric("평점", f"{rating:.1f} {stars}")

    with review_cols[1]:
        pos = review.get("positive_keywords", [])
        neg = review.get("negative_keywords", [])
        if pos:
            st.markdown("**긍정 키워드**")
            st.markdown(" ".join(f":green[{k}]" for k in pos))
        if neg:
            st.markdown("**부정 키워드**")
            st.markdown(" ".join(f":red[{k}]" for k in neg))

    with review_cols[2]:
        summary = review.get("summary_text", "")
        if summary:
            st.markdown(f"**리뷰 요약:** {summary}")


def render_comparison_table(products: list[dict]) -> pd.DataFrame:
    """제품 비교 테이블 생성 및 렌더링."""
    rows = []
    for p in products:
        if p.get("error") and p.get("product_name") == "추출 실패":
            continue
        review = p.get("review_summary", {})
        rows.append({
            "제품명": p.get("product_name", "정보 없음"),
            "브랜드": p.get("brand", "정보 없음"),
            "가격": p.get("price_display", "정보 없음"),
            "원산지": p.get("country_of_origin", "정보 없음"),
            "소재": p.get("materials", "정보 없음"),
            "옵션": ", ".join(p.get("options", [])) or "정보 없음",
            "크기": p.get("size", "정보 없음"),
            "리뷰 요약": review.get("summary_text", "정보 없음"),
            "특이사항": " / ".join(p.get("notable_features", [])) or "정보 없음",
        })

    if not rows:
        st.info("비교할 제품이 없습니다.")
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)
    return df


def render_usp_analysis(comparison: dict):
    """고유 판매 포인트 분석 결과 렌더링."""
    if not comparison:
        return

    # 시장 요약
    market_summary = comparison.get("market_summary", "")
    if market_summary:
        st.info(f"**시장 포지셔닝 요약:** {market_summary}")

    recommendation = comparison.get("recommendation", "")
    if recommendation:
        st.success(f"**기획 시사점:** {recommendation}")

    # 제품별 분석
    analyses = comparison.get("products_analysis", [])
    for analysis in analyses:
        name = analysis.get("product_name", "알 수 없음")
        with st.expander(f"**{name}** — {analysis.get('price_positioning', '')}"):
            col1, col2 = st.columns(2)
            with col1:
                usps = analysis.get("unique_selling_points", [])
                if usps:
                    st.markdown("**고유 판매 포인트 (USP)**")
                    for u in usps:
                        st.markdown(f"- {u}")

                strengths = analysis.get("strengths", [])
                if strengths:
                    st.markdown("**강점**")
                    for s in strengths:
                        st.markdown(f"- :green[{s}]")

            with col2:
                weaknesses = analysis.get("weaknesses", [])
                if weaknesses:
                    st.markdown("**약점**")
                    for w in weaknesses:
                        st.markdown(f"- :red[{w}]")


def export_comparison_to_excel(products: list[dict]) -> bytes:
    """비교 결과를 엑셀 파일로 내보내기."""
    wb = Workbook()
    ws = wb.active
    ws.title = "제품 비교"

    # 스타일
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    body_font = Font(name="맑은 고딕", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # 헤더
    headers = ["No", "제품명", "브랜드", "가격", "원산지", "소재",
               "옵션", "크기", "리뷰 요약", "특이사항", "URL"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터
    row_idx = 2
    for i, p in enumerate(products, 1):
        if p.get("error") and p.get("product_name") == "추출 실패":
            continue
        review = p.get("review_summary", {})
        review_text = review.get("summary_text", "정보 없음")
        pos = review.get("positive_keywords", [])
        neg = review.get("negative_keywords", [])
        if pos or neg:
            review_text += f"\n긍정: {', '.join(pos)}" if pos else ""
            review_text += f"\n부정: {', '.join(neg)}" if neg else ""

        values = [
            i,
            p.get("product_name", "정보 없음"),
            p.get("brand", "정보 없음"),
            p.get("price_display", "정보 없음"),
            p.get("country_of_origin", "정보 없음"),
            p.get("materials", "정보 없음"),
            ", ".join(p.get("options", [])) or "정보 없음",
            p.get("size", "정보 없음"),
            review_text,
            " / ".join(p.get("notable_features", [])) or "정보 없음",
            p.get("url", ""),
        ]
        for j, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=j, value=v)
            cell.font = body_font
            cell.alignment = wrap_align
        row_idx += 1

    # 열 너비 설정
    widths = [5, 30, 15, 15, 10, 25, 25, 20, 40, 35, 40]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + j) if j <= 26 else ""].width = w
    # openpyxl 열 문자 변환
    from openpyxl.utils import get_column_letter
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
