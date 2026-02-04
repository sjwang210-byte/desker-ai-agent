"""엑셀 파일 파싱 — 시트 감지, 헤더 추출, 데이터 파싱."""

import re
from pathlib import Path

import openpyxl
import pandas as pd


# ── 시트 감지 패턴 ──
_MONTH_PATTERN = re.compile(r"^(\d{1,2})월$")                        # '1월', '12월'
_COST_PATTERN = re.compile(r"^(\d{1,2})월\s*하자보수비\s*금액$")      # '1월 하자보수비 금액'


def detect_sheets(filepath: str | Path) -> dict:
    """엑셀 파일의 시트 목록을 분석하여 데이터/비용 시트를 식별한다.

    Returns:
        {
            "all_sheets": [시트명, ...],
            "data_sheets": [{"name": "1월", "month": 1}, ...],
            "cost_sheets": [{"name": "1월 하자보수비 금액", "month": 1}, ...],
        }
    """
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    result = {"all_sheets": wb.sheetnames, "data_sheets": [], "cost_sheets": []}
    for name in wb.sheetnames:
        m = _MONTH_PATTERN.match(name.strip())
        if m:
            result["data_sheets"].append({"name": name, "month": int(m.group(1))})
            continue
        m = _COST_PATTERN.match(name.strip())
        if m:
            result["cost_sheets"].append({"name": name, "month": int(m.group(1))})
    wb.close()
    return result


def extract_headers(filepath: str | Path, sheet_name: str) -> list[str]:
    """지정 시트의 헤더(1행) 컬럼명 리스트 반환."""
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value is not None else "")
    wb.close()
    return headers


def parse_data_sheet(filepath: str | Path, sheet_name: str,
                     column_mapping: dict[str, str]) -> list[dict]:
    """데이터 시트를 파싱하여 매핑된 컬럼 기준의 딕셔너리 리스트 반환.

    Args:
        column_mapping: {"product_group": "품목군", "product": "제품", ...}

    Returns:
        [{"row_number": 2, "product_group": "...", ...}, ...]
    """
    df = pd.read_excel(str(filepath), sheet_name=sheet_name, engine="openpyxl")

    # 역매핑: 엑셀컬럼명 → 내부키
    reverse_map = {v: k for k, v in column_mapping.items() if v in df.columns}

    records = []
    for idx, row in df.iterrows():
        rec = {"row_number": idx + 2}  # 엑셀 행 번호 (1행=헤더)
        for excel_col, internal_key in reverse_map.items():
            val = row.get(excel_col)
            if pd.isna(val):
                val = None
            elif internal_key == "amount":
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    val = None
            else:
                val = str(val).strip() if val is not None else None
            rec[internal_key] = val

        # 매핑되지 않은 컬럼도 extra로 보관
        mapped_cols = set(reverse_map.keys())
        for col in df.columns:
            if col not in mapped_cols and not pd.isna(row.get(col)):
                val = row.get(col)
                if not pd.isna(val):
                    rec[col] = str(val) if not isinstance(val, (int, float)) else val

        records.append(rec)

    return records


def parse_cost_sheet(filepath: str | Path, sheet_name: str,
                     total_column: str = "전 체") -> float | None:
    """비용 시트에서 '전 체' 컬럼의 합계 값을 추출.

    비용 시트는 보통 소수 행만 있으며, 마지막(또는 첫) 데이터 행의 합계를 반환.
    """
    df = pd.read_excel(str(filepath), sheet_name=sheet_name, engine="openpyxl")

    # 컬럼명에 공백이 있을 수 있으므로 유연하게 매칭
    target_col = None
    for col in df.columns:
        cleaned = str(col).replace(" ", "").strip()
        if cleaned == total_column.replace(" ", ""):
            target_col = col
            break

    if target_col is None:
        # 마지막 컬럼을 시도
        target_col = df.columns[-1]

    # 숫자 값 중 가장 큰 값 (합계일 가능성 높음)
    values = pd.to_numeric(df[target_col], errors="coerce").dropna()
    if values.empty:
        return None

    return float(values.max())


def get_year_month_from_filename(filename: str) -> str | None:
    """파일명에서 연-월 정보 추출. 예: '하자보수비(브랜드)_26년01월 1.xlsx' → '2026-01'"""
    m = re.search(r"(\d{2})년\s*(\d{1,2})월", filename)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        full_year = 2000 + year if year < 100 else year
        return f"{full_year}-{month:02d}"
    return None
